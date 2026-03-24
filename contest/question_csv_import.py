"""
Bulk import Question rows from CSV or XLSX (no round ids in file).

Pick the target Round in the admin; the file only needs question data.

Required headers (case-insensitive; spaces → underscores):
  category, question_text, option_a, option_b, option_c, option_d, correct_answer

Accepted aliases (if canonical column missing):
  question → question_text
  marks, mark → points
  correct, correct_question, answer → correct_answer

Optional:
  difficulty, points, order, is_active

Any round_id column is ignored.

XLSX: requires ``openpyxl`` (pip install openpyxl). First sheet, first row = headers.
"""

from __future__ import annotations

import csv
import io
from typing import Any, Dict, List, Sequence, Tuple

from django.db import transaction
from django.db.models import Max

from .models import Question, Round

REQUIRED_FIELDS = [
    "category",
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_answer",
]

HEADER_ALIASES = {
    "question": "question_text",
    "marks": "points",
    "mark": "points",
    "correct": "correct_answer",
    "correct_question": "correct_answer",
    "answer": "correct_answer",
}

VALID_CATEGORY = {"quantitative", "logical", "verbal"}
VALID_DIFFICULTY = {"easy", "medium", "hard"}
VALID_ANSWER = {"A", "B", "C", "D"}

TEMPLATE_HEADERS = [
    "category",
    "question_text",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "correct_answer",
    "difficulty",
    "points",
    "order",
    "is_active",
]

EXAMPLE_ROW = [
    "quantitative",
    "What is 15 + 27?",
    "40",
    "42",
    "41",
    "43",
    "B",
    "easy",
    "4",
    "1",
    "true",
]


def _norm_header(name: str) -> str:
    if name is None:
        return ""
    return str(name).strip().lower().replace(" ", "_").replace("-", "_")


def _parse_bool(val: Any) -> bool:
    if val is None or str(val).strip() == "":
        return True
    s = str(val).strip().lower()
    if s in ("1", "true", "yes", "y", "on"):
        return True
    if s in ("0", "false", "no", "n", "off"):
        return False
    return True


def _truncate(s: str, max_len: int) -> str:
    s = (s or "").strip()
    return s[:max_len] if len(s) > max_len else s


def _cell_to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _build_field_map(fieldnames: Sequence[Optional[str]]) -> Dict[str, str]:
    """normalized_key -> original column name (as used in row dicts)."""
    fm = {}
    for h in fieldnames:
        if h is None:
            continue
        hs = str(h).strip()
        if not hs:
            continue
        fm[_norm_header(hs)] = hs
    for alias, canonical in HEADER_ALIASES.items():
        if alias in fm and canonical not in fm:
            fm[canonical] = fm[alias]
    return fm


def _import_questions_from_rows(
    target_round: Round,
    field_map: Dict[str, str],
    rows: List[Dict[str, Any]],
) -> Tuple[int, List[str], List[str]]:
    """Shared import logic. ``rows`` are dicts keyed by original header strings."""
    errors: List[str] = []
    warnings: List[str] = []

    missing = [f for f in REQUIRED_FIELDS if f not in field_map]
    if missing:
        return 0, [f"Missing required column(s): {', '.join(missing)}"], []

    if not rows:
        return 0, ["No data rows after the header."], []

    round_id = target_round.pk
    mx = Question.objects.filter(round_id=round_id).aggregate(m=Max("order"))["m"]
    next_order = (mx or 0) + 1

    created = 0
    with transaction.atomic():
        for i, raw in enumerate(rows, start=2):
            row_num = i

            def cell(key: str) -> str:
                if key not in field_map:
                    return ""
                col = field_map[key]
                return _cell_to_str(raw.get(col))

            if not any(cell(k) for k in REQUIRED_FIELDS):
                continue

            cat = cell("category").lower()
            if cat not in VALID_CATEGORY:
                errors.append(
                    f"Row {row_num}: invalid category {cat!r} (use quantitative, logical, verbal)."
                )
                continue

            qtext = cell("question_text")
            if not qtext:
                errors.append(f"Row {row_num}: question_text is empty.")
                continue

            oa = _truncate(cell("option_a"), 500)
            ob = _truncate(cell("option_b"), 500)
            oc = _truncate(cell("option_c"), 500)
            od = _truncate(cell("option_d"), 500)
            if not all([oa, ob, oc, od]):
                errors.append(f"Row {row_num}: all four options must be non-empty.")
                continue

            ans = cell("correct_answer").upper()
            if ans not in VALID_ANSWER:
                errors.append(
                    f"Row {row_num}: correct_answer must be A, B, C, or D (got {ans!r})."
                )
                continue

            diff = cell("difficulty").lower() or "medium"
            if diff not in VALID_DIFFICULTY:
                diff = "medium"
                warnings.append(f"Row {row_num}: invalid difficulty, using 'medium'.")

            pts_s = cell("points")
            if pts_s:
                try:
                    points = int(float(pts_s))
                except ValueError:
                    points = 4
                    warnings.append(f"Row {row_num}: invalid points, using 4.")
            else:
                points = 4

            ord_s = cell("order")
            if ord_s:
                try:
                    order = int(float(ord_s))
                except ValueError:
                    order = next_order
                    next_order += 1
                    warnings.append(f"Row {row_num}: invalid order, auto-assigned.")
            else:
                order = next_order
                next_order += 1

            if "is_active" in field_map:
                active = _parse_bool(raw.get(field_map["is_active"]))
            else:
                active = True

            Question.objects.create(
                round=target_round,
                round_number=target_round.order_number,
                category=cat,
                question_text=qtext,
                option_a=oa,
                option_b=ob,
                option_c=oc,
                option_d=od,
                correct_answer=ans,
                difficulty=diff,
                points=points,
                order=order,
                is_active=active,
            )
            created += 1

    return created, errors, warnings


def import_questions_from_csv_text(
    text: str,
    target_round: Round,
) -> Tuple[int, List[str], List[str]]:
    if text.startswith("\ufeff"):
        text = text[1:]

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return 0, ["CSV has no header row."], []

    field_map = _build_field_map(reader.fieldnames)
    rows = [dict(r) for r in reader]
    return _import_questions_from_rows(target_round, field_map, rows)


def import_questions_from_xlsx_bytes(
    data: bytes,
    target_round: Round,
) -> Tuple[int, List[str], List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return (
            0,
            [
                "XLSX import requires the 'openpyxl' package. "
                "Install it with: pip install openpyxl"
            ],
            [],
        )

    bio = io.BytesIO(data)
    try:
        wb = load_workbook(bio, read_only=True, data_only=True)
    except Exception as e:
        return 0, [f"Could not read Excel file: {e}"], []

    ws = wb.active
    row_values = list(ws.iter_rows(values_only=True))
    wb.close()

    if not row_values:
        return 0, ["Excel sheet is empty."], []

    header_row = row_values[0]
    headers = [
        str(h).strip() if h is not None and str(h).strip() != "" else ""
        for h in header_row
    ]
    field_map = _build_field_map(headers)

    rows: List[Dict[str, Any]] = []
    for rv in row_values[1:]:
        lst = list(rv) if rv else []
        d: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            d[h] = lst[i] if i < len(lst) else None
        if any(v is not None and str(v).strip() != "" for v in d.values()):
            rows.append(d)

    return _import_questions_from_rows(target_round, field_map, rows)


def csv_template_content() -> str:
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_HEADERS)
    w.writerow(EXAMPLE_ROW)
    return buf.getvalue()


def xlsx_template_bytes() -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(TEMPLATE_HEADERS)
    ws.append(EXAMPLE_ROW)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
